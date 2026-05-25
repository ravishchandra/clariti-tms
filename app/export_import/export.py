"""XLSX export writer for Phase 5 Excel round-trip.

The format produced here is the v1 wire contract documented in
``docs/07-excel-roundtrip.md``. Every detail (column order, locked-vs-editable
cells, dropdown values, ``_meta`` hash) is read back by the import side; do
not change shapes here without bumping :data:`app.export_import.schema.SCHEMA_VERSION`.

Design notes
------------
* **Read-only.** Export never writes to ``translations`` or any other table.
  We deliberately don't take a transaction or write through ``app.mt.api`` —
  a SELECT-only path doesn't cross a module boundary in any harmful sense.
* **Synchronous.** The HTTP endpoint streams the workbook bytes back in the
  same request. The spec calls for async polling for very large exports;
  ship sync first, upgrade when reviewer feedback says we need it (TODO).
* **NFC normalisation** on every string cell — keeps Excel's hidden Unicode
  recomposition from looking like reviewer-introduced changes at import.
"""

from __future__ import annotations

import hashlib
import io
import logging
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import case, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.export_import.schema import (
    COL_CURRENT_TRANSLATION,
    COL_INTERNAL_ID,
    COL_REVIEWER_ACTION,
    COLUMN_INDEX,
    COLUMN_ORDER,
    COLUMN_WIDTHS,
    EDITABLE_COLUMNS,
    META_FIELDS,
    META_STATUS_FILTER_ALL,
    REVIEWER_ACTIONS,
    SCHEMA_VERSION,
    SHEET_PROTECTION_PASSWORD,
    STATUS_COLOR_MAP,
)
from app.models import Key, Project, Translation

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public data shapes
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExportRow:
    """One row in a locale tab. Source-of-truth field names match
    :data:`app.export_import.schema.COLUMN_ORDER` 1:1.
    """

    key: str
    source_text: str
    description: str | None
    component: str | None
    screen: str | None
    max_length: int | None
    placeholders: list[str]
    risk_class: str
    current_translation: str | None
    mt_suggestion: str | None
    status: str
    internal_id: uuid.UUID

    def normalised(self) -> ExportRow:
        """Return a copy with all string fields NFC-normalised.

        Excel normalises some characters silently on edit; if we wrote NFD
        and reviewers saved the file, the import would see "edits" that
        weren't really edits. Pinning everything to NFC on the way out
        lines up with the same normalisation on import.
        """
        return ExportRow(
            key=_nfc(self.key),
            source_text=_nfc(self.source_text),
            description=_nfc(self.description) if self.description is not None else None,
            component=_nfc(self.component) if self.component is not None else None,
            screen=_nfc(self.screen) if self.screen is not None else None,
            max_length=self.max_length,
            placeholders=[_nfc(p) for p in self.placeholders],
            risk_class=self.risk_class,
            current_translation=_nfc(self.current_translation) if self.current_translation is not None else None,
            mt_suggestion=_nfc(self.mt_suggestion) if self.mt_suggestion is not None else None,
            status=self.status,
            internal_id=self.internal_id,
        )


@dataclass(frozen=True)
class ExportRequest:
    """Inputs to :func:`build_workbook` — everything needed to deterministically
    materialise an XLSX. ``rows_by_locale`` preserves caller-chosen ordering
    (``dict`` insertion order); tabs land in the same order.
    """

    project_id: uuid.UUID
    project_slug: str
    status_filter: str | None  # None == no filter; export everything
    exported_by_email: str | None  # None for CLI / system exports
    exported_by_user_id: uuid.UUID | None
    export_timestamp: datetime
    rows_by_locale: dict[str, list[ExportRow]] = field(default_factory=dict)


def _nfc(s: str) -> str:
    """Normalise a string to Unicode NFC. ``None`` passes through callers."""
    return unicodedata.normalize("NFC", s)


# ---------------------------------------------------------------------------
# DB read path
# ---------------------------------------------------------------------------


async def fetch_export_rows(
    db: AsyncSession,
    project_id: uuid.UUID,
    locales: list[str],
    status_filter: str | None,
    sample_size: int | None = None,
) -> dict[str, list[ExportRow]]:
    """Fetch translation rows for the given project + locales + optional status.

    Returns a ``{locale: [ExportRow, ...]}`` map. A locale present in
    ``locales`` but with zero matching rows is still represented (empty list).
    Rows within each locale are ordered by ``(component, screen, key)`` so
    reviewers see related strings grouped together — the same order the
    review UI uses.

    ``sample_size`` is the bootstrap-export knob (docs/15 F-3): when set, the
    rows are re-ordered by descending risk class first (so reviewers see the
    most consequential strings) and then limited to that count *per locale*.
    Typical use is ``sample_size=50`` with ``locales=[one_locale]``.

    Read-only. No writes, no commits. The DB session may be in any state.
    """
    out: dict[str, list[ExportRow]] = {locale: [] for locale in locales}
    if not locales:
        return out

    # Risk-class priority for bootstrap sampling. Higher number = sampled first.
    # Matches docs/05 risk-class enum (auto_publish < standard < high_risk <
    # human_only) so the bootstrap reviewer sees the highest-stakes strings.
    risk_priority = case(
        (Key.risk_class == "human_only", 4),
        (Key.risk_class == "high_risk", 3),
        (Key.risk_class == "standard", 2),
        (Key.risk_class == "auto_publish", 1),
        else_=0,
    )

    if sample_size is None:
        # Default ordering: grouped by component/screen/key — matches the
        # review UI's natural grouping and is what the existing full-export
        # flow has always emitted.
        stmt = (
            select(Translation, Key)
            .join(Key, Translation.key_id == Key.id)
            .where(
                Key.project_id == project_id,
                Key.is_active.is_(True),
                Translation.locale.in_(locales),
            )
            .order_by(Translation.locale, Key.component, Key.screen, Key.key)
        )
        if status_filter is not None:
            stmt = stmt.where(Translation.status == status_filter)
        result = await db.execute(stmt)
    else:
        # Bootstrap-sample path: one query per locale so the LIMIT applies
        # per-locale (otherwise a global LIMIT would overweight whichever
        # locale comes first alphabetically). N small queries, N ≤ 5 typical.
        async def _sample(locale: str) -> list:
            stmt = (
                select(Translation, Key)
                .join(Key, Translation.key_id == Key.id)
                .where(
                    Key.project_id == project_id,
                    Key.is_active.is_(True),
                    Translation.locale == locale,
                )
                .order_by(risk_priority.desc(), Key.created_at, Key.key)
                .limit(sample_size)
            )
            if status_filter is not None:
                stmt = stmt.where(Translation.status == status_filter)
            res = await db.execute(stmt)
            return res.all()

        # Awaiting sequentially keeps the same AsyncSession safe.
        all_rows: list = []
        for locale in locales:
            all_rows.extend(await _sample(locale))

        class _PlainResult:
            def __init__(self, rows: list) -> None:
                self._rows = rows

            def all(self) -> list:
                return self._rows

        result = _PlainResult(all_rows)  # type: ignore[assignment]
    for translation, key in result.all():
        placeholders_raw = key.placeholders or []
        if isinstance(placeholders_raw, list):
            placeholders = [str(p) for p in placeholders_raw]
        else:
            # Defensive: JSONB column could in principle contain an object;
            # render to a stable string so a malformed key doesn't crash export.
            placeholders = [str(placeholders_raw)]

        row = ExportRow(
            key=key.key,
            source_text=key.source_text,
            description=key.description,
            component=key.component,
            screen=key.screen,
            max_length=key.max_length,
            placeholders=placeholders,
            risk_class=key.risk_class,
            current_translation=translation.value,
            mt_suggestion=translation.mt_value,
            status=translation.status,
            internal_id=translation.id,
        )
        out.setdefault(translation.locale, []).append(row)

    return out


async def fetch_project_for_export(
    db: AsyncSession,
    project_slug: str,
) -> Project | None:
    """Resolve a project slug -> Project (or None). Convenience for CLI."""
    return await db.scalar(select(Project).where(Project.slug == project_slug))


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------


def build_workbook(request: ExportRequest) -> Workbook:
    """Build a Workbook from an :class:`ExportRequest`. Pure / deterministic
    given the same input. Does not touch the DB or any I/O.
    """
    wb = Workbook()
    # Workbook() creates a default blank sheet; drop it — we add tabs explicitly
    # so we can control ordering.
    default = wb.active
    if default is not None:
        wb.remove(default)

    locales = list(request.rows_by_locale.keys())
    row_counts: dict[str, int] = {}

    for locale in locales:
        rows = [r.normalised() for r in request.rows_by_locale[locale]]
        _build_locale_tab(wb, locale, rows)
        row_counts[locale] = len(rows)

    _build_meta_tab(wb, request, row_counts)
    return wb


def _build_locale_tab(wb: Workbook, locale: str, rows: list[ExportRow]) -> None:
    ws: Worksheet = wb.create_sheet(title=locale)

    # ---- Header row -------------------------------------------------------
    header_font = Font(bold=True)
    for idx, col_name in enumerate(COLUMN_ORDER, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        # Header is locked along with the rest of columns 1-11 and 15 by
        # default; openpyxl Cell.protection defaults to locked=True.

    # ---- Data rows --------------------------------------------------------
    for row_idx, row in enumerate(rows, start=2):
        _write_row(ws, row_idx, row)

    # ---- Column widths ----------------------------------------------------
    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---- Hide _internal_id column ----------------------------------------
    # docs/07 line 40: hidden + locked. Column letter for index 15 == "O".
    ws.column_dimensions[get_column_letter(COLUMN_INDEX[COL_INTERNAL_ID])].hidden = True

    # ---- Frozen header ----------------------------------------------------
    # "A2" freezes row 1 (everything above the cursor cell stays visible).
    ws.freeze_panes = "A2"

    # ---- Reviewer-action dropdown ----------------------------------------
    # The dropdown is bound to all data cells in column 12. We use an inline
    # list rather than referencing a range so the workbook has no hidden
    # "lookup" sheet to leak structure on a reviewer share.
    dv_formula = '"' + ",".join(REVIEWER_ACTIONS) + '"'
    dv = DataValidation(
        type="list",
        formula1=dv_formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid reviewer action",
        error=f"Allowed: {', '.join(REVIEWER_ACTIONS)} (or blank).",
    )
    # Apply to col 12 from row 2 down to the last data row. If the sheet is
    # empty we still bind one row so reviewers who paste new data see the
    # dropdown — paste isn't really supported but the failure mode is
    # better with the dropdown present.
    last_row = max(2, len(rows) + 1)
    col_letter = get_column_letter(COLUMN_INDEX[COL_REVIEWER_ACTION])
    dv.add(f"{col_letter}2:{col_letter}{last_row}")
    ws.add_data_validation(dv)

    # ---- Sheet protection -------------------------------------------------
    # protection.sheet=True activates the lock-flags on every cell. Editable
    # cells are explicitly unlocked above (in _write_row); everything else
    # honours the default locked=True. Password is UX, not security
    # (see schema.SHEET_PROTECTION_PASSWORD docstring).
    ws.protection.sheet = True
    ws.protection.password = SHEET_PROTECTION_PASSWORD
    # Allow the common review actions even with the sheet protected. Without
    # these flags the reviewer can't even select cells.
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


def _write_row(ws: Worksheet, row_idx: int, row: ExportRow) -> None:
    """Write one data row. Text columns get @ text format to defeat Excel's
    auto-coercion; editable columns get unlocked; the whole row's
    ``current_translation`` cell gets a colour wash by status.
    """
    values: list[Any] = [
        row.key,
        row.source_text,
        row.description or "",
        row.component or "",
        row.screen or "",
        row.max_length,  # may be None — Excel renders blank
        ", ".join(row.placeholders),
        row.risk_class,
        row.current_translation or "",
        row.mt_suggestion or "",
        row.status,
        "",  # reviewer_action — empty by default
        "",  # edited_translation
        "",  # notes
        str(row.internal_id),
    ]

    status_fill: PatternFill | None = None
    color_hex = STATUS_COLOR_MAP.get(row.status)
    if color_hex is not None:
        status_fill = PatternFill(fgColor=color_hex, fill_type="solid")

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)

        # Text format on every column except max_length (which is genuinely
        # numeric). This prevents Excel from "fixing" leading zeros or long
        # numeric-looking strings like phone numbers in source_text.
        if col_idx != 6:
            cell.number_format = "@"

        # Unlock the three editable columns; everything else keeps the
        # default locked=True from openpyxl. Cell.protection is itself a
        # StyleProxy that must be reassigned, not mutated.
        if col_idx in EDITABLE_COLUMNS:
            cell.protection = Protection(locked=False)
        # Wrap text in the long-string columns so reviewers can read at a
        # glance without resizing.
        if col_idx in (2, 9, 10, 13):
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Colour the current_translation cell by status. We chose the single-cell
    # variant over whole-row fill because (a) it preserves readability of
    # the locked metadata columns and (b) reviewers scan that column most.
    if status_fill is not None:
        ws.cell(row=row_idx, column=COLUMN_INDEX[COL_CURRENT_TRANSLATION]).fill = status_fill


def _build_meta_tab(
    wb: Workbook,
    request: ExportRequest,
    row_counts: dict[str, int],
) -> None:
    """Write the hidden ``_meta`` tab. Field order tracks
    :data:`META_FIELDS` so the import side can do an ordered read.
    """
    ws: Worksheet = wb.create_sheet(title="_meta")

    locales = list(request.rows_by_locale.keys())

    # Compute the export hash now so the value goes into the workbook. We
    # collect every row's internal_id, sort, and SHA-256 the concatenation
    # together with the project id + timestamp. The import side recomputes
    # this hash from the workbook contents and flags mismatch.
    export_hash = compute_export_hash(
        project_id=request.project_id,
        export_timestamp=request.export_timestamp,
        row_ids=[row.internal_id for locale in locales for row in request.rows_by_locale[locale]],
    )

    field_values: dict[str, str] = {
        "schema_version": SCHEMA_VERSION,
        "export_timestamp": _iso_utc(request.export_timestamp),
        "project_id": str(request.project_id),
        "project_slug": request.project_slug,
        "exported_by": request.exported_by_email or "",
        "exported_by_user_id": str(request.exported_by_user_id) if request.exported_by_user_id else "",
        "status_filter": request.status_filter if request.status_filter is not None else META_STATUS_FILTER_ALL,
        "locales": ", ".join(locales),
        "row_counts": ", ".join(f"{locale}: {row_counts.get(locale, 0)}" for locale in locales),
        "export_hash": export_hash,
        "notes": "Generated by ClaritiTMS Phase 5 exporter (v0.x)",
    }

    # Header row for readability when reviewers / support do open the tab.
    ws.cell(row=1, column=1, value="Field").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Value").font = Font(bold=True)

    for idx, field_name in enumerate(META_FIELDS, start=2):
        ws.cell(row=idx, column=1, value=field_name)
        v_cell = ws.cell(row=idx, column=2, value=field_values[field_name])
        v_cell.number_format = "@"  # don't let Excel reformat the hash etc.

    ws.column_dimensions["A"].width = 24.0
    ws.column_dimensions["B"].width = 60.0

    # The meta tab is reference material, not a review surface. Hide it.
    # Reviewers can still right-click -> Unhide; that's fine, we just don't
    # want it to be the first thing they see.
    ws.sheet_state = "hidden"


def compute_export_hash(
    project_id: uuid.UUID,
    export_timestamp: datetime,
    row_ids: list[uuid.UUID],
) -> str:
    """Deterministic export hash.

    Inputs: project_id, export_timestamp (rendered as ISO-8601 UTC), and the
    sorted list of translation UUIDs included in the export. Output: hex
    SHA-256.

    The import side will recompute this from the workbook and compare. Any
    structural tamper — a missing row, a swapped internal_id, a forged
    timestamp — flips the hash.
    """
    sorted_ids = sorted(str(rid) for rid in row_ids)
    payload = "\n".join([str(project_id), _iso_utc(export_timestamp), *sorted_ids])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _iso_utc(ts: datetime) -> str:
    """Render a datetime as ISO-8601 with a trailing ``Z`` (UTC).

    Naive datetimes are treated as UTC — caller's responsibility. We always
    produce a Z-suffixed string so the hash is stable across timezones.
    """
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=UTC)
    else:
        ts = ts.astimezone(UTC)
    return ts.strftime("%Y-%m-%dT%H:%M:%SZ")


# ---------------------------------------------------------------------------
# Top-level convenience: produce the raw bytes the HTTP / CLI layers stream.
# ---------------------------------------------------------------------------


def export_to_bytes(request: ExportRequest) -> bytes:
    """Materialise a Workbook from ``request`` and return the .xlsx bytes."""
    wb = build_workbook(request)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Filename helper — used by both the CLI and the HTTP endpoint so the naming
# convention from docs/07-excel-roundtrip.md line 15 is honoured exactly once.
# ---------------------------------------------------------------------------


def build_filename(
    project_slug: str,
    locales: list[str],
    status_filter: str | None,
    export_timestamp: datetime,
) -> str:
    """``{slug}_{locales}_{YYYY-MM-DD}_{status}.xlsx``.

    Multi-locale exports collapse the locale segment to ``multi`` to keep
    filenames sane on shells.
    """
    date_part = export_timestamp.strftime("%Y-%m-%d")
    if len(locales) == 1:
        locale_part = locales[0]
    else:
        locale_part = "multi"
    status_part = (status_filter or "all").replace("_", "-")
    return f"{project_slug}_{locale_part}_{date_part}_{status_part}.xlsx"
