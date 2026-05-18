"""XLSX reader for the Excel round-trip import flow.

Reads a workbook, validates the ``_meta`` tab and column order, normalizes
every cell (NFC, ``\\r\\n`` -> ``\\n``, strip), and yields
:class:`ParsedImportRow` instances per locale tab.

The parser is *strict*: any structural deviation (missing tab, wrong column
order, unsupported schema version, ``.xlsm``, CSV-disguised file) raises
:class:`ImportParseError` so the caller can surface a clear validation
failure to the user. Per-row validation errors (placeholder mismatch, max
length, etc.) are *not* raised here — they belong to :mod:`validate`.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.export_import import schema as xschema
from app.export_import.excel_traps import (
    ACCEPTED_EXTENSIONS,
    is_zip_signature,
    normalize_cell_text,
    normalize_int,
)

# Size of the magic-byte prefix we sniff for `is_zip_signature`. 4 bytes is
# enough for the ZIP local-file-header check.
_MAGIC_PREFIX_BYTES: Final[int] = 8


class ImportParseError(Exception):
    """Raised for structural failures: missing meta, wrong columns, etc."""


@dataclass
class ParsedImportRow:
    """One row from a locale tab, after Excel-trap normalization.

    Every text field has been NFC-normalized and `\\r\\n` collapsed. Blank
    cells become ``None`` (or ``""`` for cells the reviewer touched and left
    empty — we can't reliably distinguish, so we treat ``""`` as "blank").
    """

    # Row index in the workbook, 1-based, for error messages.
    row_index: int
    locale: str

    # Identity. `internal_id` is the FK back to translations.id.
    internal_id: uuid.UUID
    key: str

    # Reference (locked, informational).
    source_text: str
    description: str | None
    component: str | None
    screen: str | None
    max_length: int | None
    placeholders_raw: str | None
    risk_class: str | None
    current_translation: str | None
    mt_suggestion: str | None
    status: str | None

    # Editable fields.
    reviewer_action: str | None  # raw from XLSX: yes/no/edit/needs_more_context or blank
    edited_translation: str | None
    notes: str | None

    # Hidden — needed for conflict detection.
    source_hash_at_export: str | None


@dataclass
class ParsedMeta:
    """`_meta` tab fields. Whatever the exporter wrote, normalized."""

    schema_version: str
    project_id: uuid.UUID
    project_slug: str | None
    export_timestamp: datetime | None
    exported_by: str | None
    exported_by_user_id: uuid.UUID | None
    status_filter: str | None
    locales: list[str]
    row_counts: dict[str, int]
    export_hash: str | None
    notes: str | None


@dataclass
class ParsedImport:
    """The complete parsed workbook — meta + rows grouped by locale."""

    meta: ParsedMeta
    rows_by_locale: dict[str, list[ParsedImportRow]] = field(default_factory=dict)

    @property
    def total_rows(self) -> int:
        return sum(len(rows) for rows in self.rows_by_locale.values())


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------


def parse_xlsx_path(path: str | Path) -> ParsedImport:
    """Parse a workbook from a filesystem path.

    Rejects non-``.xlsx`` files at the extension layer; rejects CSV-
    disguised-as-xlsx files via a magic-byte sniff. Then delegates to
    :func:`parse_xlsx_bytes`.
    """
    p = Path(path)
    if p.suffix.lower() not in ACCEPTED_EXTENSIONS:
        raise ImportParseError(
            f"Unsupported file extension {p.suffix!r}. Only .xlsx is accepted (no .xlsm/macros, no CSV)."
        )
    data = p.read_bytes()
    return parse_xlsx_bytes(data, filename=p.name)


def parse_xlsx_bytes(data: bytes, filename: str | None = None) -> ParsedImport:
    """Parse a workbook from in-memory bytes.

    Use this from the HTTP layer where the upload arrives as
    :class:`fastapi.UploadFile`. Rejects empty payloads and non-ZIP magic
    bytes (catches CSV-renamed-to-xlsx).
    """
    if not data:
        raise ImportParseError("Empty file.")
    if not is_zip_signature(data[:_MAGIC_PREFIX_BYTES]):
        raise ImportParseError(
            "File does not look like a real .xlsx (no ZIP magic bytes). "
            "Reject CSV / plain-text uploads disguised as xlsx."
        )

    # Filename-extension guard for callers that pass filename through. The
    # path entry point checks extension before reading bytes; the bytes entry
    # point repeats the check defensively for callers that bypass the path
    # API.
    if filename is not None:
        suffix = Path(filename).suffix.lower()
        if suffix and suffix not in ACCEPTED_EXTENSIONS:
            raise ImportParseError(
                f"Unsupported file extension {suffix!r}. Only .xlsx is accepted (no .xlsm/macros, no CSV)."
            )

    import io

    try:
        # data_only=True: read computed values rather than formulae. read_only
        # would be faster but we need named-sheet access and the typical Phase
        # 5 file is < 10k rows, so default mode is fine.
        wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    except Exception as exc:  # openpyxl raises a variety of error subclasses
        raise ImportParseError(f"Could not open workbook: {exc}") from exc

    return _parse_workbook(wb)


# ---------------------------------------------------------------------------
# Internal — workbook walking
# ---------------------------------------------------------------------------


def _parse_workbook(wb: Workbook) -> ParsedImport:
    meta = _parse_meta(wb)

    # Every declared locale must have a sheet of the same name.
    rows_by_locale: dict[str, list[ParsedImportRow]] = {}
    for locale in meta.locales:
        if locale not in wb.sheetnames:
            raise ImportParseError(
                f"_meta declared locale {locale!r} but no sheet named {locale!r} exists. Sheet names: {wb.sheetnames!r}"
            )
        rows_by_locale[locale] = _parse_locale_sheet(wb[locale], locale)

    return ParsedImport(meta=meta, rows_by_locale=rows_by_locale)


def _parse_meta(wb: Workbook) -> ParsedMeta:
    if xschema.META_SHEET_NAME not in wb.sheetnames:
        raise ImportParseError(f"Workbook is missing the required {xschema.META_SHEET_NAME!r} tab.")
    ws = wb[xschema.META_SHEET_NAME]

    # _meta is a two-column key/value sheet. Header row is "Field" / "Value"
    # per docs/07:60-74, but accept either with or without a header.
    raw: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        key, value = row[0], row[1]
        if key is None:
            continue
        k = normalize_cell_text(key)
        if k is None or k.lower() in {"field", ""}:
            continue
        v = normalize_cell_text(value) if value is not None else None
        raw[k] = v or ""

    schema_version = raw.get(xschema.META_KEY_SCHEMA_VERSION, "")
    if schema_version not in xschema.SUPPORTED_SCHEMA_VERSIONS:
        raise ImportParseError(
            f"Unsupported schema_version {schema_version!r}. "
            f"This importer supports: {sorted(xschema.SUPPORTED_SCHEMA_VERSIONS)}. "
            "See docs/07-excel-roundtrip.md 'Schema versioning policy'."
        )

    project_id_raw = raw.get(xschema.META_KEY_PROJECT_ID, "")
    try:
        project_id = uuid.UUID(project_id_raw)
    except (ValueError, TypeError) as exc:
        raise ImportParseError(
            f"_meta.{xschema.META_KEY_PROJECT_ID} is missing or not a UUID: {project_id_raw!r}"
        ) from exc

    # exported_by_user_id is optional (some exports don't have a user).
    user_id: uuid.UUID | None = None
    user_id_raw = raw.get(xschema.META_KEY_EXPORTED_BY_USER_ID, "")
    if user_id_raw:
        try:
            user_id = uuid.UUID(user_id_raw)
        except (ValueError, TypeError):
            user_id = None  # ignore — non-critical metadata

    export_ts: datetime | None = None
    export_ts_raw = raw.get(xschema.META_KEY_EXPORT_TIMESTAMP, "")
    if export_ts_raw:
        export_ts = _parse_iso_datetime(export_ts_raw)

    locales_raw = raw.get(xschema.META_KEY_LOCALES, "")
    locales = [s.strip() for s in locales_raw.split(",") if s.strip()]
    if not locales:
        raise ImportParseError(f"_meta.{xschema.META_KEY_LOCALES} is empty — workbook has no locale tabs declared.")

    row_counts_raw = raw.get(xschema.META_KEY_ROW_COUNTS, "")
    row_counts = _parse_row_counts(row_counts_raw)

    return ParsedMeta(
        schema_version=schema_version,
        project_id=project_id,
        project_slug=raw.get(xschema.META_KEY_PROJECT_SLUG) or None,
        export_timestamp=export_ts,
        exported_by=raw.get(xschema.META_KEY_EXPORTED_BY) or None,
        exported_by_user_id=user_id,
        status_filter=raw.get(xschema.META_KEY_STATUS_FILTER) or None,
        locales=locales,
        row_counts=row_counts,
        export_hash=raw.get(xschema.META_KEY_EXPORT_HASH) or None,
        notes=raw.get(xschema.META_KEY_NOTES) or None,
    )


def _parse_iso_datetime(value: str) -> datetime | None:
    """Best-effort ISO-8601 parse. Returns ``None`` on failure (non-fatal)."""
    # `Z` suffix -> +00:00 for fromisoformat.
    cleaned = value.strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(cleaned)
    except ValueError:
        return None


def _parse_row_counts(value: str) -> dict[str, int]:
    """`fr-FR: 247, de-DE: 198` -> ``{'fr-FR': 247, 'de-DE': 198}``."""
    result: dict[str, int] = {}
    for segment in value.split(","):
        if ":" not in segment:
            continue
        loc, count = segment.split(":", 1)
        try:
            result[loc.strip()] = int(count.strip())
        except ValueError:
            continue
    return result


def _parse_locale_sheet(ws: Worksheet, locale: str) -> list[ParsedImportRow]:
    """Read a locale tab into a list of :class:`ParsedImportRow`.

    Row 1 is the header. We validate the header order exactly matches
    :data:`xschema.LOCALE_TAB_COLUMNS`; mismatches raise
    :class:`ImportParseError` because re-import requires the contract.
    """
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportParseError(f"Locale sheet {locale!r} is empty (no header row).") from None

    header = [normalize_cell_text(c) for c in header_row]
    _validate_header(header, locale)

    # Build name -> index map for the columns we read.
    col_index: dict[str, int] = {}
    for idx, name in enumerate(header):
        if name is not None:
            col_index[name] = idx

    rows: list[ParsedImportRow] = []
    for row_offset, raw_row in enumerate(rows_iter, start=2):  # row_offset = 1-based row in the sheet, header is row 1
        # Skip totally empty rows — reviewers sometimes leave blanks at the end.
        if raw_row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw_row):
            continue

        # Internal id — required identity. If absent, the row is a reviewer-
        # added freeform that we can't safely apply. Raise so the user fixes
        # the file (the alternative — silently skipping — risks dropping a
        # genuine review action).
        internal_id_raw = _cell(raw_row, col_index, xschema.COL_INTERNAL_ID)
        if internal_id_raw is None or not internal_id_raw.strip():
            raise ImportParseError(
                f"Row {row_offset} on sheet {locale!r} is missing _internal_id. "
                "Did the reviewer insert a row by hand? Re-import requires the "
                "hidden _internal_id column to identify the translation."
            )
        try:
            internal_id = uuid.UUID(internal_id_raw.strip())
        except (ValueError, TypeError) as exc:
            raise ImportParseError(
                f"Row {row_offset} on sheet {locale!r}: _internal_id is not a valid UUID: {internal_id_raw!r}"
            ) from exc

        key = _cell(raw_row, col_index, xschema.COL_KEY) or ""

        # reviewer_action is read in lowercase to be forgiving; the doc spec
        # is lower-case (yes/no/edit/needs_more_context).
        ra = _cell(raw_row, col_index, xschema.COL_REVIEWER_ACTION)
        if ra is not None:
            ra = ra.lower()
            if ra == "":
                ra = None

        max_len = normalize_int(_raw_cell(raw_row, col_index, xschema.COL_MAX_LENGTH))

        rows.append(
            ParsedImportRow(
                row_index=row_offset,
                locale=locale,
                internal_id=internal_id,
                key=key,
                source_text=_cell(raw_row, col_index, xschema.COL_SOURCE_TEXT) or "",
                description=_cell(raw_row, col_index, xschema.COL_DESCRIPTION),
                component=_cell(raw_row, col_index, xschema.COL_COMPONENT),
                screen=_cell(raw_row, col_index, xschema.COL_SCREEN),
                max_length=max_len,
                placeholders_raw=_cell(raw_row, col_index, xschema.COL_PLACEHOLDERS),
                risk_class=_cell(raw_row, col_index, xschema.COL_RISK_CLASS),
                current_translation=_cell(raw_row, col_index, xschema.COL_CURRENT_TRANSLATION),
                mt_suggestion=_cell(raw_row, col_index, xschema.COL_MT_SUGGESTION),
                status=_cell(raw_row, col_index, xschema.COL_STATUS),
                reviewer_action=ra,
                edited_translation=_cell(raw_row, col_index, xschema.COL_EDITED_TRANSLATION),
                notes=_cell(raw_row, col_index, xschema.COL_NOTES),
                source_hash_at_export=_cell(raw_row, col_index, xschema.COL_SOURCE_HASH_AT_EXPORT),
            )
        )
    return rows


def _validate_header(header: list[str | None], locale: str) -> None:
    expected = list(xschema.LOCALE_TAB_COLUMNS)
    # We compare position-by-position; extra trailing columns past the
    # expected length are permitted (a future v2 might add columns we don't
    # know about — but v1 import-time, we ignore them rather than fail).
    got = header[: len(expected)]
    # Pad got with None if shorter.
    while len(got) < len(expected):
        got.append(None)

    if got != expected:
        mismatches = [
            f"col {i + 1}: expected {expected[i]!r}, got {got[i]!r}"
            for i in range(len(expected))
            if got[i] != expected[i]
        ]
        raise ImportParseError(
            f"Locale sheet {locale!r} has the wrong column order. Expected {expected!r}. Mismatches: {mismatches}"
        )


def _cell(row: tuple, col_index: dict[str, int], col_name: str) -> str | None:
    """Read a cell by column name, NFC/strip/newline-normalized. ``None`` if blank."""
    idx = col_index.get(col_name)
    if idx is None or idx >= len(row):
        return None
    return normalize_cell_text(row[idx])


def _raw_cell(row: tuple, col_index: dict[str, int], col_name: str) -> object:
    """Read a cell by column name without any normalization. ``None`` if absent."""
    idx = col_index.get(col_name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


__all__ = [
    "ImportParseError",
    "ParsedImport",
    "ParsedImportRow",
    "ParsedMeta",
    "parse_xlsx_bytes",
    "parse_xlsx_path",
]
