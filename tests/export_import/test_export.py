"""Tests for the Phase 5 XLSX exporter.

These tests intentionally avoid the DB layer: they assemble :class:`ExportRow`
fixtures in-memory and pass them through :func:`build_workbook` /
:func:`export_to_bytes`. The DB-touching paths (``fetch_export_rows``,
``/api/v1/exports``) are tested in the wider integration suite that has
Postgres available.

What we verify here:
  * Column order matches the v1 schema exactly.
  * Frozen header, hidden _internal_id, dropdown on col 12.
  * Text format ("@") prevents Excel from rewriting leading-zero strings.
  * Color coding fills the current_translation cell per status.
  * Multi-locale exports produce one tab per locale + a hidden _meta tab.
  * The _meta tab fields are present and in the right shape.
  * export_hash is deterministic for the same input set (set membership, not
    list ordering inside a locale).
  * Empty result sets still emit a valid workbook with the tab structure.
  * NFC normalisation runs on string fields on the way out.
"""

from __future__ import annotations

import io
import unicodedata
import uuid
from datetime import UTC, datetime

import openpyxl
import pytest

from app.export_import.export import (
    ExportRequest,
    ExportRow,
    build_filename,
    build_workbook,
    compute_export_hash,
    export_to_bytes,
)
from app.export_import.schema import (
    COL_CURRENT_TRANSLATION,
    COL_INTERNAL_ID,
    COLUMN_INDEX,
    COLUMN_ORDER,
    META_FIELDS,
    REVIEWER_ACTIONS,
    SCHEMA_VERSION,
    STATUS_COLOR_MAP,
)

# ---------------------------------------------------------------------------
# Fixtures: build small in-memory row sets so we exercise the workbook writer
# without touching Postgres. Each fixture pins UUIDs so the export_hash tests
# can compare against a deterministic expectation.
# ---------------------------------------------------------------------------


def _row(
    key: str,
    *,
    internal_id: uuid.UUID,
    source_text: str = "Hello",
    current_translation: str | None = "Bonjour",
    mt_suggestion: str | None = "Bonjour",
    status: str = "needs_review",
    component: str | None = "Greeting",
    screen: str | None = "home",
    max_length: int | None = 50,
    placeholders: list[str] | None = None,
    risk_class: str = "standard",
    description: str | None = "Welcome banner",
) -> ExportRow:
    return ExportRow(
        key=key,
        source_text=source_text,
        description=description,
        component=component,
        screen=screen,
        max_length=max_length,
        placeholders=list(placeholders) if placeholders is not None else [],
        risk_class=risk_class,
        current_translation=current_translation,
        mt_suggestion=mt_suggestion,
        status=status,
        internal_id=internal_id,
    )


_PROJECT_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")
_TIMESTAMP = datetime(2026, 5, 17, 14, 32, 0, tzinfo=UTC)


def _request(rows_by_locale: dict[str, list[ExportRow]], status_filter: str | None = "needs_review") -> ExportRequest:
    return ExportRequest(
        project_id=_PROJECT_ID,
        project_slug="clariti-web",
        status_filter=status_filter,
        exported_by_email="ravish@clariti.app",
        exported_by_user_id=uuid.UUID("22222222-2222-2222-2222-222222222222"),
        export_timestamp=_TIMESTAMP,
        rows_by_locale=rows_by_locale,
    )


def _load(buf: bytes) -> openpyxl.Workbook:
    """Round-trip the bytes through openpyxl so we read what's actually on disk."""
    return openpyxl.load_workbook(io.BytesIO(buf))


# ---------------------------------------------------------------------------
# Single-locale export — covers the headline acceptance criteria.
# ---------------------------------------------------------------------------


class TestSingleLocaleExport:
    def test_workbook_has_locale_tab_plus_meta(self) -> None:
        rows = [_row("greeting.hello", internal_id=uuid.uuid4())]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        assert "fr-FR" in wb.sheetnames
        assert "_meta" in wb.sheetnames

    def test_header_columns_exact_order(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": [_row("k", internal_id=uuid.uuid4())]})))
        ws = wb["fr-FR"]
        header = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMN_ORDER))]
        # docs/07-excel-roundtrip.md lines 22-40 — exact match.
        assert tuple(header) == COLUMN_ORDER

    def test_frozen_header_row(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": [_row("k", internal_id=uuid.uuid4())]})))
        ws = wb["fr-FR"]
        # freeze_panes="A2" pins row 1.
        assert ws.freeze_panes == "A2"

    def test_internal_id_column_is_hidden(self) -> None:
        rows = [_row("k", internal_id=uuid.UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"))]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        # Col 15 == "O".
        assert ws.column_dimensions["O"].hidden is True

    def test_internal_id_value_round_trips_as_uuid_string(self) -> None:
        rid = uuid.UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")
        rows = [_row("k", internal_id=rid)]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        assert ws.cell(row=2, column=COLUMN_INDEX[COL_INTERNAL_ID]).value == str(rid)

    def test_reviewer_action_dropdown_present(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": [_row("k", internal_id=uuid.uuid4())]})))
        ws = wb["fr-FR"]
        # openpyxl exposes data validations on the worksheet.
        dvs = list(ws.data_validations.dataValidation)
        # Exactly one DV, bound to the reviewer_action column (L == col 12).
        assert len(dvs) == 1
        dv = dvs[0]
        assert dv.type == "list"
        # The formula is the comma-joined dropdown values wrapped in quotes.
        for action in REVIEWER_ACTIONS:
            assert action in dv.formula1
        sqref = str(dv.sqref)
        # Should reference column L (12).
        assert "L" in sqref

    def test_status_color_fills_current_translation_cell(self) -> None:
        rid = uuid.uuid4()
        rows = [_row("k", internal_id=rid, status="needs_review")]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        cell = ws.cell(row=2, column=COLUMN_INDEX[COL_CURRENT_TRANSLATION])
        # openpyxl normalises ARGB to upper-case hex.
        expected = STATUS_COLOR_MAP["needs_review"].upper()
        assert cell.fill.fgColor.rgb == expected
        assert cell.fill.patternType == "solid"

    def test_status_with_no_color_map_does_not_set_fill(self) -> None:
        rid = uuid.uuid4()
        rows = [_row("k", internal_id=rid, status="draft")]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        cell = ws.cell(row=2, column=COLUMN_INDEX[COL_CURRENT_TRANSLATION])
        # "draft" is not in STATUS_COLOR_MAP — no fill applied.
        assert cell.fill.patternType in (None, "none")

    def test_text_columns_have_text_number_format(self) -> None:
        rows = [_row("k", internal_id=uuid.uuid4())]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        # All columns except max_length (col 6) should be "@" (text).
        for col in (1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15):
            cell = ws.cell(row=2, column=col)
            assert cell.number_format == "@", f"col {col} should be '@', got {cell.number_format!r}"

    def test_leading_zero_strings_survive_round_trip(self) -> None:
        # The Excel trap: ``001234`` becomes ``1234`` if stored as a number.
        # Text format ("@") + writing it as a Python str must preserve it.
        rows = [_row("phone", internal_id=uuid.uuid4(), source_text="001234")]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        val = ws.cell(row=2, column=2).value
        assert val == "001234"
        assert isinstance(val, str)

    def test_sheet_is_protected(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": [_row("k", internal_id=uuid.uuid4())]})))
        ws = wb["fr-FR"]
        # Sheet protection on; password set to the constant.
        assert ws.protection.sheet is True

    def test_editable_columns_are_unlocked(self) -> None:
        rows = [_row("k", internal_id=uuid.uuid4())]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        # Columns 12, 13, 14 must be unlocked for reviewer edits.
        for col in (12, 13, 14):
            cell = ws.cell(row=2, column=col)
            assert cell.protection.locked is False, f"col {col} must be unlocked"

    def test_locked_columns_stay_locked(self) -> None:
        rows = [_row("k", internal_id=uuid.uuid4())]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        # Columns 1-11 and 15 must be locked (the openpyxl default).
        for col in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 15):
            cell = ws.cell(row=2, column=col)
            assert cell.protection.locked is True, f"col {col} should be locked"


# ---------------------------------------------------------------------------
# Multi-locale exports
# ---------------------------------------------------------------------------


class TestMultiLocaleExport:
    def test_one_tab_per_locale_plus_meta(self) -> None:
        rows_fr = [_row("k1", internal_id=uuid.uuid4())]
        rows_de = [
            _row("k1", internal_id=uuid.uuid4(), current_translation="Hallo"),
            _row("k2", internal_id=uuid.uuid4(), current_translation="Welt"),
        ]
        wb = _load(export_to_bytes(_request({"fr-FR": rows_fr, "de-DE": rows_de})))
        assert "fr-FR" in wb.sheetnames
        assert "de-DE" in wb.sheetnames
        assert "_meta" in wb.sheetnames
        assert len(wb.sheetnames) == 3

    def test_tab_order_follows_request_order(self) -> None:
        rows = {"de-DE": [], "fr-FR": [], "es-ES": []}
        wb = _load(export_to_bytes(_request(rows)))
        # _meta is always last; locale tabs preserve the dict order we gave.
        assert wb.sheetnames == ["de-DE", "fr-FR", "es-ES", "_meta"]


# ---------------------------------------------------------------------------
# _meta tab
# ---------------------------------------------------------------------------


class TestMetaTab:
    def test_meta_tab_is_hidden(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": []})))
        ws = wb["_meta"]
        assert ws.sheet_state == "hidden"

    def test_meta_has_all_required_fields(self) -> None:
        rows = [_row("k", internal_id=uuid.uuid4())]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["_meta"]
        # Skip header row 1; META_FIELDS occupy rows 2..N.
        observed = {ws.cell(row=i + 2, column=1).value for i in range(len(META_FIELDS))}
        assert observed == set(META_FIELDS)

    def test_meta_values_have_right_shapes(self) -> None:
        rows = [_row("k", internal_id=uuid.uuid4())]
        wb = _load(export_to_bytes(_request({"fr-FR": rows, "de-DE": rows})))
        ws = wb["_meta"]
        # Build a {field: value} dict from rows 2..N.
        meta = {ws.cell(row=i + 2, column=1).value: ws.cell(row=i + 2, column=2).value for i in range(len(META_FIELDS))}
        assert meta["schema_version"] == SCHEMA_VERSION
        assert meta["project_id"] == str(_PROJECT_ID)
        assert meta["project_slug"] == "clariti-web"
        assert meta["status_filter"] == "needs_review"
        # locales is a CSV; both locales must appear.
        assert "fr-FR" in meta["locales"]
        assert "de-DE" in meta["locales"]
        # row_counts is "<locale>: <n>" CSV.
        assert "fr-FR: 1" in meta["row_counts"]
        assert "de-DE: 1" in meta["row_counts"]
        # Timestamp ends with Z (UTC stamp).
        assert meta["export_timestamp"].endswith("Z")
        # Hash is a 64-char lower-hex SHA-256.
        assert len(meta["export_hash"]) == 64
        assert all(c in "0123456789abcdef" for c in meta["export_hash"])

    def test_status_filter_all_when_none(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": []}, status_filter=None)))
        ws = wb["_meta"]
        meta = {ws.cell(row=i + 2, column=1).value: ws.cell(row=i + 2, column=2).value for i in range(len(META_FIELDS))}
        assert meta["status_filter"] == "all"


# ---------------------------------------------------------------------------
# export_hash determinism
# ---------------------------------------------------------------------------


class TestExportHash:
    def test_same_input_yields_same_hash(self) -> None:
        rid_a = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001")
        rid_b = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000002")
        h1 = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_a, rid_b])
        h2 = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_a, rid_b])
        assert h1 == h2

    def test_hash_is_order_independent_on_row_ids(self) -> None:
        rid_a = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001")
        rid_b = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000002")
        # The contract says "sorted row ids" — we sort inside the function.
        h1 = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_a, rid_b])
        h2 = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_b, rid_a])
        assert h1 == h2

    def test_hash_changes_when_a_row_changes(self) -> None:
        rid_a = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001")
        rid_b = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000002")
        rid_c = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000003")
        h_ab = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_a, rid_b])
        h_ac = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_a, rid_c])
        assert h_ab != h_ac

    def test_hash_changes_when_timestamp_changes(self) -> None:
        rid_a = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001")
        t1 = datetime(2026, 5, 17, 14, 32, 0, tzinfo=UTC)
        t2 = datetime(2026, 5, 17, 14, 33, 0, tzinfo=UTC)
        h1 = compute_export_hash(_PROJECT_ID, t1, [rid_a])
        h2 = compute_export_hash(_PROJECT_ID, t2, [rid_a])
        assert h1 != h2

    def test_workbook_meta_hash_matches_recomputed_hash(self) -> None:
        # End-to-end: the hash written to the workbook equals what we get
        # back by recomputing from the same inputs. This is the import-side
        # check the agent will run.
        rid_a = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001")
        rid_b = uuid.UUID("aaaaaaaa-0000-0000-0000-000000000002")
        rows = [
            _row("k1", internal_id=rid_a),
            _row("k2", internal_id=rid_b),
        ]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["_meta"]
        meta = {ws.cell(row=i + 2, column=1).value: ws.cell(row=i + 2, column=2).value for i in range(len(META_FIELDS))}
        expected = compute_export_hash(_PROJECT_ID, _TIMESTAMP, [rid_a, rid_b])
        assert meta["export_hash"] == expected


# ---------------------------------------------------------------------------
# Empty result set
# ---------------------------------------------------------------------------


class TestEmptyResults:
    def test_empty_locale_still_emits_tab(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": []})))
        assert "fr-FR" in wb.sheetnames
        ws = wb["fr-FR"]
        # Header row present even with no data.
        header = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMN_ORDER))]
        assert tuple(header) == COLUMN_ORDER
        # Row 2 has no key value.
        assert ws.cell(row=2, column=1).value is None

    def test_empty_export_meta_row_counts_zero(self) -> None:
        wb = _load(export_to_bytes(_request({"fr-FR": [], "de-DE": []})))
        ws = wb["_meta"]
        meta = {ws.cell(row=i + 2, column=1).value: ws.cell(row=i + 2, column=2).value for i in range(len(META_FIELDS))}
        assert "fr-FR: 0" in meta["row_counts"]
        assert "de-DE: 0" in meta["row_counts"]


# ---------------------------------------------------------------------------
# Unicode normalisation
# ---------------------------------------------------------------------------


class TestUnicodeNormalisation:
    def test_nfd_input_normalises_to_nfc_on_write(self) -> None:
        # "café" composed two ways: NFC ("é" = U+00E9) and NFD ("e" + U+0301).
        nfd = "café"
        assert unicodedata.is_normalized("NFD", nfd)
        assert not unicodedata.is_normalized("NFC", nfd)

        rows = [_row("greet", internal_id=uuid.uuid4(), source_text=nfd, current_translation=nfd)]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        assert ws.cell(row=2, column=2).value == "café"
        assert unicodedata.is_normalized("NFC", ws.cell(row=2, column=2).value)


# ---------------------------------------------------------------------------
# Filename helper — covered here because both the CLI and HTTP path call it
# and the convention is part of the docs/07 spec (line 15).
# ---------------------------------------------------------------------------


class TestBuildFilename:
    def test_single_locale_filename(self) -> None:
        name = build_filename("clariti-web", ["fr-FR"], "needs_review", _TIMESTAMP)
        assert name == "clariti-web_fr-FR_2026-05-17_needs-review.xlsx"

    def test_multi_locale_collapses_to_multi(self) -> None:
        name = build_filename("clariti-web", ["fr-FR", "de-DE"], "needs_review", _TIMESTAMP)
        assert name == "clariti-web_multi_2026-05-17_needs-review.xlsx"

    def test_no_status_filter_renders_all(self) -> None:
        name = build_filename("clariti-web", ["fr-FR"], None, _TIMESTAMP)
        assert name == "clariti-web_fr-FR_2026-05-17_all.xlsx"


# ---------------------------------------------------------------------------
# Workbook builder pure-function check — used by both endpoints.
# ---------------------------------------------------------------------------


class TestBuildWorkbook:
    def test_build_workbook_returns_workbook_instance(self) -> None:
        wb = build_workbook(_request({"fr-FR": []}))
        assert isinstance(wb, openpyxl.Workbook)

    @pytest.mark.parametrize(
        "n_rows",
        [0, 1, 5, 50],
    )
    def test_data_row_count_matches_input(self, n_rows: int) -> None:
        rows = [_row(f"key-{i}", internal_id=uuid.uuid4()) for i in range(n_rows)]
        wb = _load(export_to_bytes(_request({"fr-FR": rows})))
        ws = wb["fr-FR"]
        # Find the highest row with content in column 1 (the key).
        last_key_row = 1
        for r in range(2, n_rows + 2):
            if ws.cell(row=r, column=1).value is not None:
                last_key_row = r
        assert last_key_row == (n_rows + 1 if n_rows else 1)
