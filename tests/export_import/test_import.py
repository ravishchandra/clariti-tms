"""Tests for the Excel-import side of Phase 5.

Two sets of tests live here:

1. **Pure** unit tests for the parser, Excel-traps helpers, and the
   summary-building logic. No DB. These run anywhere ``pytest`` runs.
2. **Integration** tests for preview / commit / rollback that drive a real
   Postgres at ``DATABASE_URL`` (the convention used by
   ``tests/mt/conftest.py``).

Pure tests run unconditionally. Integration tests are guarded by
``DATABASE_URL`` being set; in CI the harness exports
``DATABASE_URL=postgresql+asyncpg://tms:tms@localhost:5432/tms_ph5_import``
and the tests proceed.
"""

from __future__ import annotations

import hashlib
import io
import os
import uuid
from collections.abc import AsyncIterator
from datetime import UTC, datetime, timedelta
from typing import Any

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

os.environ.setdefault("DEBUG", "true")

from app.export_import import schema as xschema  # noqa: E402
from app.export_import.commit import (  # noqa: E402
    JOB_STATUS_COMMITTED,
    JOB_STATUS_PREVIEW,
    JOB_STATUS_ROLLED_BACK,
    ProjectMismatchError,
    RollbackExpiredError,
    RollbackNotAllowedError,
    commit_import,
    preview_import,
    rollback_import,
)
from app.export_import.excel_traps import (  # noqa: E402
    is_zip_signature,
    normalize_cell_text,
    normalize_int,
)
from app.export_import.parse import ImportParseError, parse_xlsx_bytes  # noqa: E402
from app.models import (  # noqa: E402
    Key,
    LocaleConfig,
    Organization,
    Project,
    Repository,
    Translation,
    TranslationStatus,
    User,
)

# ===========================================================================
# Helpers — workbook builder
# ===========================================================================


def _sha256(s: str) -> str:
    return hashlib.sha256(s.encode()).hexdigest()


def build_workbook(
    *,
    project_id: uuid.UUID,
    project_slug: str = "test-proj",
    schema_version: str = "v1",
    export_timestamp: datetime | None = None,
    locales: dict[str, list[dict[str, Any]]] | None = None,
) -> bytes:
    """Build a v1-compatible workbook in-memory and return its bytes.

    ``locales`` maps locale -> list of row dicts. Each row dict supplies a
    subset of column values; defaults fill the rest. A reviewer_action of
    ``None`` writes a blank cell.
    """
    if locales is None:
        locales = {}
    if export_timestamp is None:
        # Default to "now + 1h" so the workbook claims to have been exported
        # *after* any DB writes the test just performed. The conflict detector
        # treats translation.updated_at > export_timestamp as
        # "translation_modified_externally"; without this fudge, the test
        # fixture's just-committed translation would always look stale.
        export_timestamp = (datetime.now(tz=UTC) + timedelta(hours=1)).replace(microsecond=0)

    wb = Workbook()
    # Remove default sheet.
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Locale tabs.
    for locale_name, rows in locales.items():
        ws = wb.create_sheet(locale_name)
        ws.append(list(xschema.LOCALE_TAB_COLUMNS))
        for row in rows:
            cells: list[Any] = []
            for col in xschema.LOCALE_TAB_COLUMNS:
                value = row.get(col)
                if col == xschema.COL_INTERNAL_ID and value is not None:
                    value = str(value)
                cells.append(value)
            ws.append(cells)

    # _meta tab.
    meta = wb.create_sheet(xschema.META_SHEET_NAME)
    meta.append(["Field", "Value"])
    meta.append([xschema.META_KEY_SCHEMA_VERSION, schema_version])
    meta.append([xschema.META_KEY_EXPORT_TIMESTAMP, export_timestamp.isoformat()])
    meta.append([xschema.META_KEY_PROJECT_ID, str(project_id)])
    meta.append([xschema.META_KEY_PROJECT_SLUG, project_slug])
    meta.append([xschema.META_KEY_LOCALES, ",".join(locales.keys())])
    row_counts = ", ".join(f"{loc}: {len(rs)}" for loc, rs in locales.items())
    meta.append([xschema.META_KEY_ROW_COUNTS, row_counts])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Pure unit tests — Excel traps
# ===========================================================================


class TestExcelTraps:
    def test_zip_signature_real_xlsx(self) -> None:
        # PK\x03\x04 is the ZIP local-file-header magic.
        assert is_zip_signature(b"PK\x03\x04xxxx") is True

    def test_zip_signature_csv_disguised(self) -> None:
        # CSV-disguised-as-xlsx -> reject.
        assert is_zip_signature(b"key,value\n1,2") is False

    def test_zip_signature_short_buffer(self) -> None:
        assert is_zip_signature(b"PK") is False

    def test_normalize_text_strips(self) -> None:
        assert normalize_cell_text("  hello  ") == "hello"

    def test_normalize_text_crlf_to_lf(self) -> None:
        assert normalize_cell_text("a\r\nb") == "a\nb"
        assert normalize_cell_text("a\rb") == "a\nb"

    def test_normalize_text_nfc(self) -> None:
        # NFD "é" = e + combining acute, NFC = single precomposed.
        decomposed = "café"
        composed = "café"
        assert normalize_cell_text(decomposed) == composed

    def test_normalize_text_preserves_leading_zero_string(self) -> None:
        # Strings stay strings — leading zeros preserved.
        assert normalize_cell_text("001234") == "001234"

    def test_normalize_text_none(self) -> None:
        assert normalize_cell_text(None) is None

    def test_normalize_int_blank(self) -> None:
        assert normalize_int("") is None
        assert normalize_int(None) is None

    def test_normalize_int_whole_float(self) -> None:
        assert normalize_int(5.0) == 5

    def test_normalize_int_fractional_float_rejected(self) -> None:
        assert normalize_int(5.5) is None


# ===========================================================================
# Pure unit tests — parser
# ===========================================================================


class TestParser:
    def test_rejects_empty(self) -> None:
        with pytest.raises(ImportParseError):
            parse_xlsx_bytes(b"")

    def test_rejects_csv_disguised(self) -> None:
        with pytest.raises(ImportParseError, match="ZIP magic"):
            parse_xlsx_bytes(b"key,value\n1,2", filename="fake.xlsx")

    def test_rejects_xlsm_extension(self) -> None:
        # Need real ZIP magic to get past the magic-byte check.
        data = build_workbook(project_id=uuid.uuid4(), locales={"fr-FR": []})
        with pytest.raises(ImportParseError, match=".xlsm"):
            parse_xlsx_bytes(data, filename="macros.xlsm")

    def test_rejects_schema_v0(self) -> None:
        data = build_workbook(project_id=uuid.uuid4(), schema_version="v0", locales={"fr-FR": []})
        with pytest.raises(ImportParseError, match="schema_version"):
            parse_xlsx_bytes(data)

    def test_rejects_schema_v2(self) -> None:
        data = build_workbook(project_id=uuid.uuid4(), schema_version="v2", locales={"fr-FR": []})
        with pytest.raises(ImportParseError, match="schema_version"):
            parse_xlsx_bytes(data)

    def test_happy_path(self) -> None:
        pid = uuid.uuid4()
        tid = uuid.uuid4()
        data = build_workbook(
            project_id=pid,
            locales={
                "fr-FR": [
                    {
                        xschema.COL_KEY: "hello",
                        xschema.COL_SOURCE_TEXT: "Hello",
                        xschema.COL_MT_SUGGESTION: "Bonjour",
                        xschema.COL_REVIEWER_ACTION: "yes",
                        xschema.COL_INTERNAL_ID: tid,
                        xschema.COL_SOURCE_HASH_AT_EXPORT: _sha256("Hello"),
                    }
                ]
            },
        )
        parsed = parse_xlsx_bytes(data)
        assert parsed.meta.project_id == pid
        assert parsed.meta.schema_version == "v1"
        assert "fr-FR" in parsed.rows_by_locale
        rows = parsed.rows_by_locale["fr-FR"]
        assert len(rows) == 1
        assert rows[0].internal_id == tid
        assert rows[0].reviewer_action == "yes"
        assert rows[0].mt_suggestion == "Bonjour"

    def test_nfc_normalization_through_parse(self) -> None:
        pid = uuid.uuid4()
        tid = uuid.uuid4()
        decomposed = "café"
        composed = "café"
        data = build_workbook(
            project_id=pid,
            locales={
                "fr-FR": [
                    {
                        xschema.COL_KEY: "café",
                        xschema.COL_SOURCE_TEXT: decomposed,
                        xschema.COL_MT_SUGGESTION: decomposed,
                        xschema.COL_REVIEWER_ACTION: "yes",
                        xschema.COL_INTERNAL_ID: tid,
                    }
                ]
            },
        )
        parsed = parse_xlsx_bytes(data)
        row = parsed.rows_by_locale["fr-FR"][0]
        assert row.source_text == composed
        assert row.mt_suggestion == composed

    def test_crlf_normalized(self) -> None:
        pid = uuid.uuid4()
        tid = uuid.uuid4()
        data = build_workbook(
            project_id=pid,
            locales={
                "fr-FR": [
                    {
                        xschema.COL_KEY: "k",
                        xschema.COL_SOURCE_TEXT: "Line 1\r\nLine 2",
                        xschema.COL_MT_SUGGESTION: "Ligne 1\r\nLigne 2",
                        xschema.COL_REVIEWER_ACTION: "yes",
                        xschema.COL_INTERNAL_ID: tid,
                    }
                ]
            },
        )
        parsed = parse_xlsx_bytes(data)
        row = parsed.rows_by_locale["fr-FR"][0]
        assert "\r" not in row.source_text
        assert "\n" in row.source_text
        assert row.mt_suggestion is not None
        assert "\r" not in row.mt_suggestion

    def test_leading_zero_preserved(self) -> None:
        pid = uuid.uuid4()
        tid = uuid.uuid4()
        data = build_workbook(
            project_id=pid,
            locales={
                "fr-FR": [
                    {
                        xschema.COL_KEY: "k",
                        xschema.COL_SOURCE_TEXT: "001234",
                        xschema.COL_INTERNAL_ID: tid,
                    }
                ]
            },
        )
        parsed = parse_xlsx_bytes(data)
        assert parsed.rows_by_locale["fr-FR"][0].source_text == "001234"

    def test_missing_internal_id_raises(self) -> None:
        pid = uuid.uuid4()
        data = build_workbook(
            project_id=pid,
            locales={
                "fr-FR": [
                    {
                        xschema.COL_KEY: "k",
                        xschema.COL_SOURCE_TEXT: "Hello",
                        xschema.COL_REVIEWER_ACTION: "yes",
                        # _internal_id deliberately omitted
                    }
                ]
            },
        )
        with pytest.raises(ImportParseError, match="_internal_id"):
            parse_xlsx_bytes(data)

    def test_blank_rows_skipped(self) -> None:
        pid = uuid.uuid4()
        tid = uuid.uuid4()
        data = build_workbook(
            project_id=pid,
            locales={
                "fr-FR": [
                    {
                        xschema.COL_KEY: "k",
                        xschema.COL_SOURCE_TEXT: "Hello",
                        xschema.COL_INTERNAL_ID: tid,
                    },
                    {},  # entirely blank — should not raise
                ]
            },
        )
        parsed = parse_xlsx_bytes(data)
        # Blank row produces a row where every column is None — our skip
        # heuristic treats it as empty.
        assert len(parsed.rows_by_locale["fr-FR"]) == 1


# ===========================================================================
# Integration tests — require Postgres at $DATABASE_URL
# ===========================================================================


# Apply the asyncio session loop-scope mark only to the integration test
# class below, not module-wide. Marking module-wide would attach the asyncio
# mark to the pure unit tests above (TestExcelTraps, TestParser) and emit
# PytestWarning noise for every non-async test.
_integration_mark = pytest.mark.asyncio(loop_scope="session")


_WIPE_TABLES = (
    "translation_history",
    "mt_runs",
    "import_jobs",
    "translation_memory",
    "translations",
    "translation_batches",
    "keys",
    "component_contexts",
    "locale_configs",
    "glossary_terms",
    "repositories",
    "projects",
    "users",
    "organizations",
)


@pytest_asyncio.fixture(loop_scope="session", scope="session")
async def _import_engine() -> AsyncIterator[Any]:
    """Session-scoped async engine for the import tests.

    Named uniquely (``_import_engine``) so it doesn't collide with
    ``tests/mt/conftest.py:_engine`` when both directories run in the same
    pytest session — sharing the fixture object would close the engine
    out from under whichever module ran second.
    """
    from app.core.settings import get_settings

    settings = get_settings()
    engine = create_async_engine(settings.DATABASE_URL, echo=False, poolclass=NullPool)
    try:
        yield engine
    finally:
        await engine.dispose()


@pytest_asyncio.fixture(loop_scope="session")
async def db(_import_engine: Any) -> AsyncIterator[AsyncSession]:
    SessionMaker = async_sessionmaker(_import_engine, expire_on_commit=False)
    session = SessionMaker()
    try:
        for tbl in _WIPE_TABLES:
            try:
                await session.execute(text(f"DELETE FROM {tbl}"))
            except Exception:
                await session.rollback()
        await session.commit()
        yield session
    finally:
        try:
            await session.close()
        except Exception:  # noqa: BLE001
            pass


@pytest_asyncio.fixture(loop_scope="session")
async def fixture_setup(db: AsyncSession) -> dict[str, Any]:
    """Minimal org + project + repo + user + key + translation tuple."""
    org = Organization(name="Imp Org", slug=f"imp-{uuid.uuid4().hex[:8]}")
    db.add(org)
    await db.flush()

    user = User(organization_id=org.id, email=f"u{uuid.uuid4().hex[:8]}@x", name="U")
    db.add(user)
    await db.flush()

    project = Project(
        organization_id=org.id,
        name="P",
        slug=f"p-{uuid.uuid4().hex[:8]}",
        target_locales=["fr-FR"],
    )
    db.add(project)
    await db.flush()

    repo = Repository(
        project_id=project.id,
        name="web",
        platform="web",
        file_format="i18next",
    )
    db.add(repo)
    await db.flush()

    locale_cfg = LocaleConfig(project_id=project.id, locale="fr-FR", is_bootstrapped=True)
    db.add(locale_cfg)
    await db.flush()

    src = "Hello, {{name}}"
    key = Key(
        repository_id=repo.id,
        project_id=project.id,
        key="greeting",
        source_text=src,
        source_hash=_sha256(src),
        risk_class="standard",
        icu_shape="plain",
        has_structural_tags=False,
        placeholders=["{{name}}"],
    )
    db.add(key)
    await db.flush()

    # Build the translation in needs_review so the importer's edges are legal.
    translation = Translation(
        key_id=key.id,
        locale="fr-FR",
        status=TranslationStatus.needs_review,
        mt_value="Bonjour, {{name}}",
        value="Bonjour, {{name}}",
    )
    db.add(translation)
    await db.commit()

    return {
        "org": org,
        "user": user,
        "project": project,
        "repo": repo,
        "key": key,
        "translation": translation,
    }


def _row_for(translation: Translation, key: Key, action: str | None = None, **overrides: Any) -> dict[str, Any]:
    row: dict[str, Any] = {
        xschema.COL_KEY: key.key,
        xschema.COL_SOURCE_TEXT: key.source_text,
        xschema.COL_MT_SUGGESTION: translation.mt_value,
        xschema.COL_CURRENT_TRANSLATION: translation.value,
        xschema.COL_STATUS: translation.status,
        xschema.COL_REVIEWER_ACTION: action,
        xschema.COL_INTERNAL_ID: translation.id,
        xschema.COL_SOURCE_HASH_AT_EXPORT: key.source_hash,
    }
    row.update(overrides)
    return row


@_integration_mark
class TestPreviewCommitRollback:
    async def test_happy_path_yes_approves(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        data = build_workbook(
            project_id=project.id,
            project_slug=project.slug,
            locales={"fr-FR": [_row_for(translation, key, action="yes")]},
        )

        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        assert job.status == JOB_STATUS_PREVIEW
        assert job.dry_run_summary is not None
        assert job.dry_run_summary["counts"]["approve"] == 1

        committed = await commit_import(db=db, job_id=job.id)
        await db.commit()
        assert committed.status == JOB_STATUS_COMMITTED
        assert committed.rollback_expires_at is not None

        await db.refresh(translation)
        assert translation.status == TranslationStatus.approved
        assert translation.reviewer_action == "accept"
        assert translation.value == "Bonjour, {{name}}"  # mt_suggestion copied in

    async def test_edit_writes_edited_translation_preserving_mt_value(
        self, db: AsyncSession, fixture_setup: dict[str, Any]
    ) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]
        original_mt = translation.mt_value

        data = build_workbook(
            project_id=project.id,
            locales={
                "fr-FR": [
                    _row_for(
                        translation,
                        key,
                        action="edit",
                        **{xschema.COL_EDITED_TRANSLATION: "Salut, {{name}}"},
                    )
                ]
            },
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        await commit_import(db=db, job_id=job.id)
        await db.commit()

        await db.refresh(translation)
        assert translation.status == TranslationStatus.approved
        assert translation.reviewer_action == "edit"
        assert translation.value == "Salut, {{name}}"
        assert translation.mt_value == original_mt  # invariant preserved

    async def test_no_rejects(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        data = build_workbook(
            project_id=project.id,
            locales={"fr-FR": [_row_for(translation, key, action="no")]},
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        await commit_import(db=db, job_id=job.id)
        await db.commit()

        await db.refresh(translation)
        assert translation.status == TranslationStatus.rejected
        assert translation.reviewer_action == "reject"

    async def test_needs_more_context_with_notes(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        data = build_workbook(
            project_id=project.id,
            locales={
                "fr-FR": [
                    _row_for(
                        translation,
                        key,
                        action="needs_more_context",
                        **{xschema.COL_NOTES: "Is this formal or informal?"},
                    )
                ]
            },
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        await commit_import(db=db, job_id=job.id)
        await db.commit()

        await db.refresh(translation)
        assert translation.status == TranslationStatus.needs_more_context
        assert translation.reviewer_action == "needs_more_context"
        assert translation.reviewer_notes == "Is this formal or informal?"

    async def test_blank_action_skipped(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]
        before_status = translation.status

        data = build_workbook(
            project_id=project.id,
            locales={"fr-FR": [_row_for(translation, key, action=None)]},
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        assert job.dry_run_summary is not None
        assert job.dry_run_summary["counts"]["skip"] == 1

        await commit_import(db=db, job_id=job.id)
        await db.commit()

        await db.refresh(translation)
        assert translation.status == before_status  # unchanged

    async def test_source_changed_conflict_skipped_by_default(
        self, db: AsyncSession, fixture_setup: dict[str, Any]
    ) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]
        before_status = translation.status

        # Workbook claims a different source_hash than what's in the DB.
        data = build_workbook(
            project_id=project.id,
            locales={
                "fr-FR": [
                    _row_for(
                        translation,
                        key,
                        action="yes",
                        **{xschema.COL_SOURCE_HASH_AT_EXPORT: _sha256("DIFFERENT_SOURCE")},
                    )
                ]
            },
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        assert job.dry_run_summary is not None
        assert job.dry_run_summary["conflicts"]["source_changed"] == 1

        await commit_import(db=db, job_id=job.id)
        await db.commit()

        # Conflicted row was skipped by default — translation unchanged.
        await db.refresh(translation)
        assert translation.status == before_status

    async def test_validation_error_forces_needs_review(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        # Source has {{name}} placeholder; the edited translation drops it
        # -> placeholder mismatch -> validation error.
        data = build_workbook(
            project_id=project.id,
            locales={
                "fr-FR": [
                    _row_for(
                        translation,
                        key,
                        action="edit",
                        **{xschema.COL_EDITED_TRANSLATION: "Bonjour"},  # missing {{name}}
                    )
                ]
            },
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        assert job.dry_run_summary is not None
        assert job.dry_run_summary["validation_error_count"] == 1

        await commit_import(db=db, job_id=job.id)
        await db.commit()

        await db.refresh(translation)
        # Forced to needs_review regardless of action=edit.
        assert translation.status == TranslationStatus.needs_review

    async def test_rollback_within_window_restores(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        prior_value = translation.value
        prior_action = translation.reviewer_action
        prior_status = translation.status

        data = build_workbook(
            project_id=project.id,
            locales={"fr-FR": [_row_for(translation, key, action="yes")]},
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        await commit_import(db=db, job_id=job.id)
        await db.commit()

        await db.refresh(translation)
        assert translation.status == TranslationStatus.approved

        rolled = await rollback_import(db=db, job_id=job.id)
        await db.commit()
        assert rolled.status == JOB_STATUS_ROLLED_BACK

        await db.refresh(translation)
        assert translation.status == prior_status
        assert translation.value == prior_value
        assert translation.reviewer_action == prior_action

    async def test_rollback_after_window_rejected(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        data = build_workbook(
            project_id=project.id,
            locales={"fr-FR": [_row_for(translation, key, action="yes")]},
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        committed = await commit_import(db=db, job_id=job.id)
        # Force the rollback window into the past.
        committed.rollback_expires_at = datetime.now(tz=UTC) - timedelta(hours=1)
        await db.commit()

        with pytest.raises(RollbackExpiredError):
            await rollback_import(db=db, job_id=job.id)

    async def test_project_mismatch_rejected(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        other_project_id = uuid.uuid4()
        # Build workbook claiming a different project id.
        data = build_workbook(
            project_id=other_project_id,
            locales={"fr-FR": [_row_for(translation, key, action="yes")]},
        )
        with pytest.raises(ProjectMismatchError):
            await preview_import(
                db=db,
                project_id=project.id,
                file_bytes=data,
                filename="r.xlsx",
                uploaded_by=user.id,
            )

    async def test_commit_idempotent(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        data = build_workbook(
            project_id=project.id,
            locales={"fr-FR": [_row_for(translation, key, action="yes")]},
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()

        c1 = await commit_import(db=db, job_id=job.id)
        await db.commit()
        c2 = await commit_import(db=db, job_id=job.id)
        await db.commit()
        # Second call returns the same job without changes.
        assert c1.id == c2.id
        assert c2.status == JOB_STATUS_COMMITTED

    async def test_double_rollback_rejected(self, db: AsyncSession, fixture_setup: dict[str, Any]) -> None:
        project = fixture_setup["project"]
        translation = fixture_setup["translation"]
        key = fixture_setup["key"]
        user = fixture_setup["user"]

        data = build_workbook(
            project_id=project.id,
            locales={"fr-FR": [_row_for(translation, key, action="yes")]},
        )
        job = await preview_import(
            db=db,
            project_id=project.id,
            file_bytes=data,
            filename="r.xlsx",
            uploaded_by=user.id,
        )
        await db.commit()
        await commit_import(db=db, job_id=job.id)
        await db.commit()
        await rollback_import(db=db, job_id=job.id)
        await db.commit()

        with pytest.raises(RollbackNotAllowedError):
            await rollback_import(db=db, job_id=job.id)
